import random
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook
import openpyxl

def score(a,b):
    axg = a/b
    bxg = b/a
    ascore = 0
    bscore = 0
    atime = []
    btime = []
    for i in range(90):
        c = random.random()
        d = random.random()
        if c < axg/90:
            ascore+=1
            atime.append(i)
        if d < bxg/90:
            bscore+=1
            btime.append(i)
    #return str(ascore) + ":" + str(bscore)
    return ascore,bscore,atime,btime

#print(score(3000,2000))

def xg(a,b):
    aa = 0
    bb = 0
    for i in range(10000):
        sc = score(a,b)
        a += sc[0]
        b += sc[1]
    return aa/10000,bb/10000

def xwdl(a,b):
    w = 0
    d = 0
    l = 0
    for i in range(10000):
        sc = score(a,b)
        if sc[0] > sc[1]:
            w+=1
        elif sc[0] == sc[1]:
            d+=1
        else:
            l+=1
    s = w+d+l
    return w/s,d/s,l/s

def pointspg(a,b,p):
    sc = score(a,b)
    if sc[0] > sc[1]:
        p+=3
    elif sc[0] == sc[1]:
        p+=1
    return p

def points(a,b,c,d,e,f):
    p = 0
    for i in range(2):
        for j in range(3):
            p = pointspg(f,a,p)
        for j in range(3):
            p = pointspg(f,b,p)
        for j in range(4):
            p = pointspg(f,c,p)
        for j in range(4):
            p = pointspg(f,d,p)
        for j in range(5):
            p = pointspg(f,e,p)
    return p

def xpoints(a,b,c,d,e,f):
    p = 0
    for i in range(1000):
        p += points(a,b,c,d,e,f)
    return p/1000


'''
print(xpoints(a,b,c,d,e,a),\
    xpoints(a,b,c,d,e,b),\
    xpoints(a,b,c,d,e,c),\
    xpoints(a,b,c,d,e,d),\
    xpoints(a,b,c,d,e,e))
'''

def leagueschedule(t,n,schedule):
    random.shuffle(t)
    t19 = t.pop(n-1)
    for i in range (n-1):
        print("round"+str(i+1))
        if i%2==0:
            print(t[i].team+"-"+t19.team)
        else:
            print(t19.team+"-"+t[i].team)
        for j in range(1,n//2):
            if j%2==0:
                print(t[(j+i)%(n-1)].team+"-"+t[(n-1-j+i)%(n-1)].team)
            else:
                print(t[(n-1-j+i)%(n-1)].team+"-"+t[(j+i)%(n-1)].team)
    for i in range (n-1):
        print("round"+str(i+n))
        if i%2==1:
            print(t[i].team+"-"+t19.team)
        else:
            print(t19.team+"-"+t[i].team)
        for j in range(1,n//2):
            if j%2==1:
                print(t[(j+i)%(n-1)].team+"-"+t[(n-1-j+i)%(n-1)].team)
            else:
                print(t[(n-1-j+i)%(n-1)].team+"-"+t[(j+i)%(n-1)].team)

def leagueschedule_array(t,n,schedule):
    t19 = t.pop(n-1)
    for i in range (n-1):
        #print("round"+str(i+1))
        schedule[i*12,0]="Round"+str(i+1)
        if i%2==0:
            #print(t[i].team+"-"+t19.team)
            schedule[i*12+1,0] = t[i].team
            schedule[i*12+1,4] = t19.team
        else:
            #print(t19.team+"-"+t[i].team)
            schedule[i*12+1,0] = t19.team
            schedule[i*12+1,4] = t[i].team
        for j in range(1,n//2):
            if j%2==0:
                #print(t[(j+i)%(n-1)].team+"-"+t[(n-1-j+i)%(n-1)].team)
                schedule[i*12+j+1,0] = t[(j+i)%(n-1)].team
                schedule[i*12+j+1,4] = t[(n-1-j+i)%(n-1)].team
            else:
                #print(t[(n-1-j+i)%(n-1)].team+"-"+t[(j+i)%(n-1)].team)
                schedule[i*12+j+1,0] = t[(n-1-j+i)%(n-1)].team
                schedule[i*12+j+1,4] = t[(j+i)%(n-1)].team
    for i in range (n-1):
        #print("round"+str(i+n))
        schedule[i*12,6]="Round"+str(i+n)
        if i%2==1:
            #print(t[i].team+"-"+t19.team)
            schedule[i*12+1,6] = t[i].team
            schedule[i*12+1,10] = t19.team
        else:
            #print(t19.team+"-"+t[i].team)
            schedule[i*12+1,6] = t19.team
            schedule[i*12+1,10] = t[i].team
        for j in range(1,n//2):
            if j%2==1:
                #print(t[(j+i)%(n-1)].team+"-"+t[(n-1-j+i)%(n-1)].team)
                schedule[i*12+j+1,6] = t[(j+i)%(n-1)].team
                schedule[i*12+j+1,10] = t[(n-1-j+i)%(n-1)].team
            else:
                #print(t[(n-1-j+i)%(n-1)].team+"-"+t[(j+i)%(n-1)].team)
                schedule[i*12+j+1,6] = t[(n-1-j+i)%(n-1)].team
                schedule[i*12+j+1,10] = t[(j+i)%(n-1)].team
    t.append(t19)

class team:
    def __init__(self):
        self.team = 0
        self.power = 0
        self.win = 0
        self.draw = 0
        self.lose = 0
        self.points = 0
        self.goal = 0
        self.goalag = 0
        self.players = []
    def __lt__(self,other):
        return self.points < other.points
    class player:
        def __init__(self,name):
            self.team = 0
            self.name = name
            self.xgoals = 0
            self.xassists = 0
            self.goals = 0
            self.assists = 0
class league:
    def __init__(self):
        self.teams = []
        self.players = []

sl = league()
pl = league()
il = league()
gl = league()
fl = league()
sl2 = league()

excel_file = openpyxl.load_workbook('fb_real.xlsx')
Sheet_teams = excel_file["teams"]

def add_player(n,a,b,name,t):
    players = [0]*n
    player_name = []
    player_xgoals = []
    player_xassists = []
    for i in range(n):
        player_name.append(Sheet_teams.cell(i+a,b).value)
        player_xgoals.append(Sheet_teams.cell(i+a,b+1).value)
        player_xassists.append(Sheet_teams.cell(i+a,b+2).value)
        players[i] = team.player(player_name[i])
        players[i].team = name
        players[i].name = players[i].name + "(" + str(players[i].team) + ")"
        players[i].xgoals = player_xgoals[i]
        players[i].xassists = player_xassists[i]
        list(filter(lambda x:x.team==name,t))[0].players.append(players[i])

#spanish teams
slschedule  = np.full((227,11),"",dtype=object)
steams = []
steamspower = []
for i in range(20):
    steams.append(Sheet_teams.cell(i+3,2).value)
    steamspower.append(Sheet_teams.cell(i+3,3).value)
ts = [0]*20
for i in range(20):
    ts[i] = team()
for i in range(len(steams)):
    ts[i].team = steams[i]
    ts[i].power = steamspower[i]

add_player(14,27,26,"Barcelona",ts)
add_player(10,27,2,"Real Madrid",ts)
add_player(10,27,6,"Atlético Madrid",ts)
add_player(7,27,10,"Athletic Bilbao",ts)
add_player(9,27,14,"Villareal",ts)
add_player(6,27,18,"Real Betis",ts)
add_player(7,27,22,"Celta Vigo",ts)

sr_numbers = [1, 4, 14, 8, 9, 0, 11, 5, 17, 19, 12, 6, 13, 10, 16, 18, 2, 7, 15, 3]
new_ts = [0]*20
for i in range(20):
    new_ts[i] = ts[sr_numbers[i]]
leagueschedule_array(new_ts,20,slschedule)
np.set_printoptions(threshold=10000000)
'''data_slschedule = pd.DataFrame(slschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_slschedule.to_excel(writer, sheet_name="sl",header=False, index=False,startrow=1, startcol=13)'''

#english teams
elschedule  = np.full((227,11),"",dtype=object)
Sheet_pl = excel_file["teams"]
eteams = []
eteamspower = []
for i in range(20):
    eteams.append(Sheet_teams.cell(i+3,5).value)
    eteamspower.append(Sheet_teams.cell(i+3,6).value)
te = [0]*20
for i in range(20):
    te[i] = team()
for i in range(len(eteams)):
    te[i].team = eteams[i]
    te[i].power = eteamspower[i]

add_player(7,41,2,"Liverpool",te)
add_player(5,41,6,"Arsenal",te)
add_player(5,41,10,"Manchester City",te)
add_player(5,41,14,"Chelsea",te)
add_player(5,41,18,"Newcastle United",te)


er_numbers = [1, 5, 3, 13, 11, 9, 8, 10, 7, 12, 15, 18, 2, 0, 19, 6, 16, 17, 14, 4]
new_te = [0]*20
for i in range(20):
    new_te[i] = te[er_numbers[i]]
leagueschedule_array(new_te,20,elschedule)
np.set_printoptions(threshold=10000000)
'''data_plschedule = pd.DataFrame(plschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_plschedule.to_excel(writer, sheet_name="pl",header=False, index=False,startrow=1, startcol=13)'''

#italian teams
ilschedule  = np.full((227,11),"",dtype=object)
iteams = []
iteamspower = []
for i in range(20):
    iteams.append(Sheet_teams.cell(i+3,8).value)
    iteamspower.append(Sheet_teams.cell(i+3,9).value)
ti = [0]*20
for i in range(20):
    ti[i] = team()
for i in range(len(iteams)):
    ti[i].team = iteams[i]
    ti[i].power = iteamspower[i]

add_player(5,55,2,"Napoli",ti)
add_player(5,55,6,"Inter Milan",ti)
add_player(5,55,10,"Juventus",ti)
add_player(5,55,14,"AC Milan",ti)

ir_numbers = [18, 6, 17, 4, 9, 14, 11, 7, 13, 19, 10, 0, 3, 12, 1, 2, 8, 15, 5, 16]
new_ti = [0]*20
for i in range(20):
    new_ti[i] = ti[ir_numbers[i]]
leagueschedule_array(new_ti,20,ilschedule)
np.set_printoptions(threshold=10000000)
'''data_plschedule = pd.DataFrame(ilschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_plschedule.to_excel(writer, sheet_name="il",header=False, index=False,startrow=1, startcol=13)'''

#german teams
glschedule  = np.full((227,11),"",dtype=object)
gteams = []
gteamspower = []
for i in range(20):
    gteams.append(Sheet_teams.cell(i+3,11).value)
    gteamspower.append(Sheet_teams.cell(i+3,12).value)
tg = [0]*20
for i in range(20):
    tg[i] = team()
for i in range(len(gteams)):
    tg[i].team = gteams[i]
    tg[i].power = gteamspower[i]

add_player(5,69,2,"Bayern Munich",tg)
add_player(5,69,6,"Borussia Dortmund",tg)

gr_numbers = [7, 11, 15, 1, 9, 17, 4, 18, 13, 10, 19, 12, 14, 6, 3, 8, 0, 5, 2, 16]
new_tg = [0]*20
for i in range(20):
    new_tg[i] = tg[gr_numbers[i]]
leagueschedule_array(new_tg,20,glschedule)
np.set_printoptions(threshold=10000000)
'''data_glschedule = pd.DataFrame(glschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_glschedule.to_excel(writer, sheet_name="gl",header=False, index=False,startrow=1, startcol=13)'''

#french teams
flschedule  = np.full((227,11),"",dtype=object)
fteams = []
fteamspower = []
for i in range(20):
    fteams.append(Sheet_teams.cell(i+3,14).value)
    fteamspower.append(Sheet_teams.cell(i+3,15).value)
tf = [0]*20
for i in range(20):
    tf[i] = team()
for i in range(len(fteams)):
    tf[i].team = fteams[i]
    tf[i].power = fteamspower[i]

add_player(10,83,2,"Paris Saint-Germain",tf)
add_player(5,83,6,"Marseille",tf)

fr_numbers = [18, 14, 17, 9, 2, 11, 6, 19, 8, 1, 16, 15, 7, 10, 13, 5, 4, 12, 0, 3]
new_tf = [0]*20
for i in range(20):
    new_tf[i] = tf[fr_numbers[i]]
leagueschedule_array(new_tf,20,flschedule)
'''np.set_printoptions(threshold=10000000)
data_flschedule = pd.DataFrame(flschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_flschedule.to_excel(writer, sheet_name="fl",header=False, index=False,startrow=1, startcol=13)'''

#portugal teams
plschedule  = np.full((227,11),"",dtype=object)
pteams = []
pteamspower = []
for i in range(3):
    pteams.append(Sheet_teams.cell(i+3,17).value)
    pteamspower.append(Sheet_teams.cell(i+3,18).value)
tp = [0]*3
for i in range(3):
    tp[i] = team()
for i in range(len(pteams)):
    tp[i].team = pteams[i]
    tp[i].power = pteamspower[i]

#netherland league
nlschedule  = np.full((227,11),"",dtype=object)
nteams = []
nteamspower = []
for i in range(3):
    nteams.append(Sheet_teams.cell(i+3,20).value)
    nteamspower.append(Sheet_teams.cell(i+3,21).value)
tn = [0]*3
for i in range(3):
    tn[i] = team()
for i in range(len(nteams)):
    tn[i].team = nteams[i]
    tn[i].power = nteamspower[i]

#austria league
alschedule  = np.full((227,11),"",dtype=object)
ateams = []
ateamspower = []
for i in range(3):
    ateams.append(Sheet_teams.cell(i+3,23).value)
    ateamspower.append(Sheet_teams.cell(i+3,24).value)
ta = [0]*3
for i in range(3):
    ta[i] = team()
for i in range(len(ateams)):
    ta[i].team = ateams[i]
    ta[i].power = ateamspower[i]

#scotland league
sclschedule  = np.full((227,11),"",dtype=object)
scteams = []
scteamspower = []
for i in range(3):
    scteams.append(Sheet_teams.cell(i+3,26).value)
    scteamspower.append(Sheet_teams.cell(i+3,27).value)
tsc = [0]*3
for i in range(3):
    tsc[i] = team()
for i in range(len(scteams)):
    tsc[i].team = scteams[i]
    tsc[i].power = scteamspower[i]

#russian league
rlschedule  = np.full((227,11),"",dtype=object)
rteams = []
rteamspower = []
for i in range(3):
    rteams.append(Sheet_teams.cell(i+3,29).value)
    rteamspower.append(Sheet_teams.cell(i+3,30).value)
tr = [0]*3
for i in range(3):
    tr[i] = team()
for i in range(len(rteams)):
    tr[i].team = rteams[i]
    tr[i].power = rteamspower[i]

#belgium league
blschedule  = np.full((227,11),"",dtype=object)
bteams = []
bteamspower = []
for i in range(3):
    bteams.append(Sheet_teams.cell(i+3,32).value)
    bteamspower.append(Sheet_teams.cell(i+3,33).value)
tb = [0]*3
for i in range(3):
    tb[i] = team()
for i in range(len(bteams)):
    tb[i].team = bteams[i]
    tb[i].power = bteamspower[i]

#inv. teams
invschedule  = np.full((227,11),"",dtype=object)
invteams = []
invteamspower = []
for i in range(6):
    invteams.append(Sheet_teams.cell(i+3,35).value)
    invteamspower.append(Sheet_teams.cell(i+3,36).value)
tinv = [0]*6
for i in range(6):
    tinv[i] = team()
for i in range(len(invteams)):
    tinv[i].team = invteams[i]
    tinv[i].power = invteamspower[i]

#spanish league 2
s2lschedule  = np.full((227,11),"",dtype=object)
s2teams = []
s2teamspower = []
for i in range(20):
    s2teams.append(Sheet_teams.cell(i+3,38).value)
    s2teamspower.append(Sheet_teams.cell(i+3,39).value)
ts2 = [0]*20
for i in range(20):
    ts2[i] = team()
for i in range(len(s2teams)):
    ts2[i].team = s2teams[i]
    ts2[i].power = s2teamspower[i]

s2r_numbers = [8, 4, 18, 9, 2, 16, 11, 3, 1, 10, 7, 5, 17, 15, 6, 12, 19, 13, 0, 14]
new_ts2 = [0]*20
for i in range(20):
    new_ts2[i] = ts2[s2r_numbers[i]]
leagueschedule_array(new_ts2,20,s2lschedule)
np.set_printoptions(threshold=10000000)
# data_s2lschedule = pd.DataFrame(s2lschedule)
# with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#     data_s2lschedule.to_excel(writer, sheet_name="sl2",header=False, index=False,startrow=1, startcol=13)

data_slschedule = pd.DataFrame(slschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_slschedule.to_excel(writer, sheet_name="sl",header=False, index=False,startrow=1, startcol=13)
data_elschedule = pd.DataFrame(elschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_elschedule.to_excel(writer, sheet_name="pl",header=False, index=False,startrow=1, startcol=13)
data_ilschedule = pd.DataFrame(ilschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_ilschedule.to_excel(writer, sheet_name="il",header=False, index=False,startrow=1, startcol=13)  
data_glschedule = pd.DataFrame(glschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_glschedule.to_excel(writer, sheet_name="gl",header=False, index=False,startrow=1, startcol=13) 
data_flschedule = pd.DataFrame(flschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_flschedule.to_excel(writer, sheet_name="fl",header=False, index=False,startrow=1, startcol=13)
data_s2lschedule = pd.DataFrame(s2lschedule)
with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    data_s2lschedule.to_excel(writer, sheet_name="sl2",header=False, index=False,startrow=1, startcol=13)
for i in range(20):
    print(new_ts[i].team)
print("")
for i in range(20):
    print(new_te[i].team)
print("")
for i in range(20):
    print(new_ti[i].team)
print("")
for i in range(20):
    print(new_tg[i].team)
print("")
for i in range(20):
    print(new_tf[i].team)
print("")
for i in range(20):
    print(new_ts2[i].team)


for i in range(len(ts)):
    for j in range(len(ts[i].players)):
        sl.players.append(ts[i].players[j])

for i in range(len(te)):
    for j in range(len(te[i].players)):
        pl.players.append(te[i].players[j])

for i in range(len(ti)):
    for j in range(len(ti[i].players)):
        il.players.append(ti[i].players[j])

for i in range(len(tg)):
    for j in range(len(tg[i].players)):
        gl.players.append(tg[i].players[j])

for i in range(len(tf)):
    for j in range(len(tf[i].players)):
        fl.players.append(tf[i].players[j])


def wdl(a,ascore,b,bscore):
    if ascore>bscore:
        a.win += 1
        b.lose += 1
    elif ascore==bscore:
        a.draw += 1
        b.draw += 1
    else:
        a.lose += 1
        b.win += 1

def addtoteam(name,xg,xa,team):
    name.xgoals = xg
    name.xassists = xa
    team.players.append(name)

def find_player(s,name):
    if np.any(s==name):
        return np.where(s == name)[0][0]
    else:
        return None

def goalandassist(a,ascore,b,bscore,scorers,assisters,astime,bstime):
    ascorer = []
    aassister = []
    bscorer = []
    bassister = []
    for i in range(ascore):
        ascorer.append(str(astime[i])+"'")
        rg = random.random()
        ra = random.random()
        r = random.random()
        rgs = 0
        ras = 0
        rgl = 0
        ral = 0
        rr=0
        for j in range(len(a.players)):
            rgl += a.players[j].xgoals
            ral += a.players[j].xassists
            if rgs < rg <= rgl or (rg==0 and j==0):
                a.players[j].goals += 1
                ascorer.append(a.players[j].name)
                rr+=1
                f = find_player(scorers,a.players[j].name)
                if f != None:
                    scorers[f][1] += 1
            elif (ras < ra <= ral or (ra==0 and j==0)) and r<0.7:
                a.players[j].assists += 1
                aassister.append(a.players[j].name)
                rr+=1
                f = find_player(assisters,a.players[j].name)
                if f != None:
                    assisters[f][1] += 1
                    assisters[f][0] = a.players[j].name 
            rgs+=a.players[j].xgoals
            ras+=a.players[j].xassists
        if rr==1:
            aassister.append("-")
    for i in range(bscore):
        bscorer.append(str(bstime[i])+"'")
        rg = random.random()
        ra = random.random()
        r = random.random()
        rgs = 0
        ras = 0
        rgl = 0
        ral = 0
        rr=0
        for j in range(len(b.players)):
            rgl += b.players[j].xgoals
            ral += b.players[j].xassists
            if rgs < rg <= rgl or (rg==0 and j==0):
                b.players[j].goals += 1
                bscorer.append(b.players[j].name)
                rr+=1
                f = find_player(scorers,b.players[j].name)
                if f != None:
                    scorers[f][1] += 1
                    scorers[f][0] = b.players[j].name 
            elif (ras < ra <= ral or (ra==0 and j==0)) and r<0.7:
                b.players[j].assists += 1
                bassister.append(b.players[j].name)
                rr+=1
                f = find_player(assisters,b.players[j].name)
                if f != None:
                    assisters[f][1] += 1
                    assisters[f][0] = b.players[j].name 
            rgs+=b.players[j].xgoals
            ras+=b.players[j].xassists
        if rr==1:
            bassister.append("-")
    return ascorer,aassister,bscorer,bassister

def game(a,b,scorers,assisters):
    if a.team != "Slytherin" and b.team != "Slytherin":
        s = score(a.power+500,b.power)
        a.goal += s[0]
        b.goalag += s[0]
        a.goalag += s[1]
        b.goal += s[1]
        ga = goalandassist(a,s[0],b,s[1],scorers,assisters,s[2],s[3])
        wdl(a,s[0],b,s[1])
        strga0 = " ".join(str(s) for s in ga[0]) 
        strga1 = " ".join(str(s) for s in ga[1]) 
        strga2 = " ".join(str(s) for s in ga[2]) 
        strga3 = " ".join(str(s) for s in ga[3]) 
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: " + strga0 + " | " + strga2)
        print("assists: " + strga1 + " | " + strga3)
        print(" ")
    else:
        s = [0,0]
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: ")
        print("assists: ")
        print(" ")
    return s

def cdrgame(a,b,scorers,assisters):
    if a.team != "Slytherin" and b.team != "Slytherin":
        s = [0,0,0,0]
        while(s[0]==s[1]):
            s = score(a.power+500,b.power)
        a.goal += s[0]
        b.goalag += s[0]
        a.goalag += s[1]
        b.goal += s[1]
        ga = goalandassist(a,s[0],b,s[1],scorers,assisters,s[2],s[3])
        wdl(a,s[0],b,s[1])
        strga0 = " ".join(str(s) for s in ga[0]) 
        strga1 = " ".join(str(s) for s in ga[1]) 
        strga2 = " ".join(str(s) for s in ga[2]) 
        strga3 = " ".join(str(s) for s in ga[3]) 
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: " + strga0 + " | " + strga2)
        print("assists: " + strga1 + " | " + strga3)
        print(" ")
    else:
        s = [0,0]
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: ")
        print("assists: ")
        print(" ")
    return s

def final_game(a,b,scorers,assisters):
    if a.team != "Slytherin" and b.team != "Slytherin":
        s = score(a.power,b.power)
        a.goal += s[0]
        b.goalag += s[0]
        a.goalag += s[1]
        b.goal += s[1]
        ga = goalandassist(a,s[0],b,s[1],scorers,assisters,s[2],s[3])
        wdl(a,s[0],b,s[1])
        strga0 = " ".join(str(s) for s in ga[0]) 
        strga1 = " ".join(str(s) for s in ga[1]) 
        strga2 = " ".join(str(s) for s in ga[2]) 
        strga3 = " ".join(str(s) for s in ga[3]) 
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: " + strga0 + " | " + strga2)
        print("assists: " + strga1 + " | " + strga3)
        print(" ")
    else:
        s = [0,0]
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: ")
        print("assists: ")
        print(" ")
    return s

def mygame(a,b,scorers,assisters,ga):
    if a.team == "Slytherin" or b.team == "Slytherin":
        s = [len(ga[0]),len(ga[2])]
        a.goal += s[0]
        b.goalag += s[0]
        a.goalag += s[1]
        b.goal += s[1]
        #ga = goalandassist(a,s[0],b,s[1],scorers,assisters)
        for i in range(len(ga[0])):
            f = find_player(scorers,ga[0][i])
            if f != None:
                    scorers[f][1] += 1
        for i in range(len(ga[1])):
            f = find_player(assisters,ga[1][i])
            if f != None:
                    assisters[f][1] += 1
        for i in range(len(ga[2])):
            f = find_player(scorers,ga[2][i])
            if f != None:
                    scorers[f][1] += 1
        for i in range(len(ga[3])):
            f = find_player(assisters,ga[3][i])
            if f != None:
                    assisters[f][1] += 1
        wdl(a,s[0],b,s[1])
        strga0 = " ".join(str(s) for s in ga[0]) 
        strga1 = " ".join(str(s) for s in ga[1]) 
        strga2 = " ".join(str(s) for s in ga[2]) 
        strga3 = " ".join(str(s) for s in ga[3]) 
        print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: " + strga0 + " | " + strga2)
        print("assists: " + strga1 + " | " + strga3)
        print(" ")
    else:
        s = [0,0]
        '''print(a.team+"  "+str(s[0])+"-"+str(s[1])+"  "+b.team)
        print("goals: ")
        print("assists: ")
        print(" ")'''
    return s

def leagueresult(t):
    random.shuffle(t)
    t19 = t.pop(19)
    for i in range (19):
        #print("round"+str(i+1))
        if i%2==0:
            game(t[i],t19)
        else:
            game(t19,t[i])
        for j in range(1,10):
            if j%2==0:
                game(t[(j+i)%(19)],t[(19-j+i)%(19)])
            else:
                game(t[(19-j+i)%(19)],t[(j+i)%(19)])
    for i in range (19):
        #print("round"+str(i+20))
        if i%2==1:
            game(t[i],t19)
        else:
            game(t19,t[i])
        for j in range(1,10):
            if j%2==1:
                game(t[(j+i)%(19)],t[(19-j+i)%(19)])
            else:
                game(t[(19-j+i)%(19)],t[(j+i)%(19)])
    t.append(t19)
    for i in range(20):
        t[i].points = t[i].win*3+t[i].draw
    t_sorted = sorted(t,reverse = True) 
    print("p t w d l p gf ga gd")
    for i in range(20):
        print(i+1,t_sorted[i].team,t_sorted[i].win,\
            t_sorted[i].draw,t_sorted[i].lose,\
            t_sorted[i].points,t_sorted[i].goal,t_sorted[i].goalag,\
            t_sorted[i].goal-t_sorted[i].goalag)

def leagueround(t,n,sheet,league):
    print(sheet)
    scorers = np.full((len(league.players),3),"",dtype=object)
    assisters = np.full((len(league.players),3),"",dtype=object)
    for i in range(len(league.players)):
        scorers[i][0] = league.players[i].name
        scorers[i][1] = 0
        scorers[i][2] = i+1
        assisters[i][0] = league.players[i].name
        assisters[i][1] = 0
        assisters[i][2] = i+1
    Sheet = excel_file[sheet]
    i = n-1
    #scores = np.full((10,3),"",dtype=object)
    scores = np.full((227,11),"",dtype=object)
    for j in range(227):
        for k in range(11):
            scores[j,k] = Sheet.cell(2+j,14+k).value
    t19 = t.pop(19)
    print("Round"+str(i+1))
    if i<19:
        if i%2==0:
            s = game(t[i],t19,scorers,assisters)
            scores[1+(i%19)*12,1] += s[0]
            scores[1+(i%19)*12,2] = "-"
            scores[1+(i%19)*12,3] += s[1]
        else:
            s = game(t19,t[i],scorers,assisters)
            scores[1+(i%19)*12,1] += s[0]
            scores[1+(i%19)*12,2] = "-"
            scores[1+(i%19)*12,3] += s[1]
        for j in range(1,10):
            if j%2==0:
                s = game(t[(j+i)%(19)],t[(19-j+i)%(19)],scorers,assisters)
                scores[j+1+(i%19)*12,1] += s[0]
                scores[j+1+(i%19)*12,2] = "-"
                scores[j+1+(i%19)*12,3] += s[1]
            else:
                s = game(t[(19-j+i)%(19)],t[(j+i)%(19)],scorers,assisters)
                scores[j+1+(i%19)*12,1] += s[0]
                scores[j+1+(i%19)*12,2] = "-"
                scores[j+1+(i%19)*12,3] += s[1]
            #print("round"+str(i+20))
    else:
        if i%2==0:
            s = game(t[i%19],t19,scorers,assisters)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        else:
            s = game(t19,t[i%19],scorers,assisters)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        for j in range(1,10):
            if j%2==1:
                s = game(t[(j+i)%(19)],t[(19-j+i)%(19)],scorers,assisters)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
            else:
                s = game(t[(19-j+i)%(19)],t[(j+i)%(19)],scorers,assisters)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
    t.append(t19)
    for i in range(20):
        t[i].points = t[i].win*3+t[i].draw
    t_sorted = sorted(t,reverse = True) 
    standings = np.full((20,10),"",dtype=object)
    for j in range(20):
        standings[j,0] = t[j].team
        standings[j,1] = t[j].points
        standings[j,2] = 1
        standings[j,3] = t[j].win
        standings[j,4] = t[j].draw
        standings[j,5] = t[j].lose
        standings[j,6] = t[j].goal
        standings[j,7] = t[j].goalag
        standings[j,8] = t[j].goal - t[j].goalag
        standings[j,9] = j+1
    data_scores = pd.DataFrame(scores)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_scores.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=1, startcol=13)
    #print(standings)
    sorted_s = scorers[np.argsort(scorers[:,2])]
    sorted_a = assisters[np.argsort(assisters[:,2])]
    return standings,sorted_s,sorted_a
    '''data_standings = pd.DataFrame(standings)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_standings.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=2, startcol=2)'''

def myleagueround(t,n,sheet,league,ga):
    print(sheet)
    scorers = np.full((len(league.players),3),"",dtype=object)
    assisters = np.full((len(league.players),3),"",dtype=object)
    for i in range(len(league.players)):
        scorers[i][0] = league.players[i].name
        scorers[i][1] = 0
        scorers[i][2] = i+1
        assisters[i][0] = league.players[i].name
        assisters[i][1] = 0
        assisters[i][2] = i+1
    Sheet = excel_file[sheet]
    i = n-1
    #scores = np.full((10,3),"",dtype=object)
    scores = np.full((227,11),"",dtype=object)
    for j in range(227):
        for k in range(11):
            scores[j,k] = Sheet.cell(2+j,14+k).value
    t19 = t.pop(19)
    print("Round"+str(i+1))
    if i<19:
        if i%2==0:
            s = mygame(t[i],t19,scorers,assisters,ga)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        else:
            s = mygame(t19,t[i],scorers,assisters,ga)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        for j in range(1,10):
            if j%2==0:
                s = mygame(t[(j+i)%(19)],t[(19-j+i)%(19)],scorers,assisters,ga)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
            else:
                s = mygame(t[(19-j+i)%(19)],t[(j+i)%(19)],scorers,assisters,ga)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
            #print("round"+str(i+20))
    else:
        if i%2==0:
            s = mygame(t[i%19],t19,scorers,assisters,ga)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        else:
            s = mygame(t19,t[i%19],scorers,assisters,ga)
            scores[1+(i%19)*12,1+(i//19)*6] += s[0]
            scores[1+(i%19)*12,2+(i//19)*6] = "-"
            scores[1+(i%19)*12,3+(i//19)*6] += s[1]
        for j in range(1,10):
            if j%2==1:
                s = mygame(t[(j+i)%(19)],t[(19-j+i)%(19)],scorers,assisters,ga)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
            else:
                s = mygame(t[(19-j+i)%(19)],t[(j+i)%(19)],scorers,assisters,ga)
                scores[j+1+(i%19)*12,1+(i//19)*6] += s[0]
                scores[j+1+(i%19)*12,2+(i//19)*6] = "-"
                scores[j+1+(i%19)*12,3+(i//19)*6] += s[1]
    t.append(t19)
    for i in range(20):
        t[i].points = t[i].win*3+t[i].draw
    t_sorted = sorted(t,reverse = True) 
    standings = np.full((20,10),"",dtype=object)
    for j in range(20):
        standings[j,0] = t[j].team
        standings[j,1] = t[j].points
        standings[j,2] = 0
        standings[j,3] = t[j].win
        standings[j,4] = t[j].draw
        standings[j,5] = t[j].lose
        standings[j,6] = t[j].goal
        standings[j,7] = t[j].goalag
        standings[j,8] = t[j].goal - t[j].goalag
        standings[j,9] = j+1
    data_scores = pd.DataFrame(scores)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_scores.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=1, startcol=13)
    #print(standings)
    sorted_s = scorers[np.argsort(scorers[:,2])]
    sorted_a = assisters[np.argsort(assisters[:,2])]
    return standings,sorted_s,sorted_a

sstandings = np.full((20,10),"",dtype=object)
pstandings = np.full((20,10),"",dtype=object)
istandings = np.full((20,10),"",dtype=object)
gstandings = np.full((20,10),"",dtype=object)
fstandings = np.full((20,10),"",dtype=object)

Sheet_sl = excel_file["sl"]
Sheet_pl = excel_file["pl"]
Sheet_il = excel_file["il"]
Sheet_gl = excel_file["gl"]
Sheet_fl = excel_file["fl"]
Sheet_cl = excel_file["cl"]
Sheet_el = excel_file["el"]
Sheet_cdr = excel_file["cdr"]

def read_standings(s,sheet):
    for i in range(20):
        for j in range(10):
            s[i,j] = sheet.cell(i+3,3+j).value
    sorted_s = s[np.argsort(s[:,9])]
    return sorted_s    
    
def makestandings(s,sheet):
    Sheet = excel_file[sheet]
    new_standings = s
    print(new_standings)
    for i in range(20):
        for j in range(1,9):
            new_standings[i,j] = s[i,j] + read_standings(sstandings,Sheet)[i,j]
    sorted_new_standings = new_standings[np.lexsort((new_standings[:,6],new_standings[:,8],new_standings[:,1]))[::-1]]
    data_standings = pd.DataFrame(sorted_new_standings)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_standings.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=2, startcol=2)

def get_scorers(n,scorers,sheet):
    Sheet = excel_file[sheet]
    for i in range(n):
        scorers[i,0] = Sheet.cell(29+i,3).value
        scorers[i,1] = Sheet.cell(29+i,4).value
        scorers[i,2] = Sheet.cell(29+i,5).value
    sorted_scorers = scorers[np.argsort(scorers[:,2])]
    return sorted_scorers

def get_assisters(n,assisters,sheet):
    Sheet = excel_file[sheet]
    for i in range(n):
        assisters[i,0] = Sheet.cell(29+i,7).value
        assisters[i,1] = Sheet.cell(29+i,8).value
        assisters[i,2] = Sheet.cell(29+i,9).value
    sorted_assisters = assisters[np.argsort(assisters[:,2])]
    return sorted_assisters

def make_score_ranking(s,sheet,n):
    Sheet = excel_file[sheet]
    initscorers = np.full((n,3),"",dtype=object)
    old_s = get_scorers(n,initscorers,sheet)
    sorted_old_s = old_s[np.argsort(old_s[:,2])]
    new_s = s
    for i in range(n):
        new_s[i][1] = sorted_old_s[i][1] + s[i][1]
    sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
    data_score = pd.DataFrame(sorted_new_s)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_score.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=28, startcol=2)

def make_assist_ranking(a,sheet,n):
    Sheet = excel_file[sheet]
    initassister = np.full((n,3),"",dtype=object)
    old_a = get_assisters(n,initassister,sheet)
    sorted_old_a = old_a[np.argsort(old_a[:,2])]
    new_a = a
    for i in range(n):
        new_a[i][1] = sorted_old_a[i][1] + a[i][1]
    sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
    data_assist = pd.DataFrame(sorted_new_a)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_assist.to_excel(writer, sheet_name=sheet,header=False, index=False,startrow=28, startcol=6)

def my_leagueround(n,ga):
    a = myleagueround(new_ts,n,"sl",sl,ga)
    makestandings(a[0],"sl")
    make_score_ranking(a[1],"sl",70)
    make_assist_ranking(a[2],"sl",70)

def all_leagueround(n):
    a = leagueround(new_ts,n,"sl",sl)
    makestandings(a[0],"sl")
    make_score_ranking(a[1],"sl",70)
    make_assist_ranking(a[2],"sl",70)
    a = leagueround(new_te,n,"pl",pl)
    makestandings(a[0],"pl")
    make_score_ranking(a[1],"pl",55)
    make_assist_ranking(a[2],"pl",55)
    a = leagueround(new_ti,n,"il",il)
    makestandings(a[0],"il")
    make_score_ranking(a[1],"il",44)
    make_assist_ranking(a[2],"il",44)
    a = leagueround(new_tg,n,"gl",gl)
    makestandings(a[0],"gl")
    make_score_ranking(a[1],"gl",22)
    make_assist_ranking(a[2],"gl",22)
    a = leagueround(new_tf,n,"fl",fl)
    makestandings(a[0],"fl")
    make_score_ranking(a[1],"fl",22)
    make_assist_ranking(a[2],"fl",22)
    a = leagueround(new_ts2,n,"sl2",sl2)
    makestandings(a[0],"sl2")


'''a = list(filter(lambda x:x.team==0,t))[0]
for i in range(len(a.players)):
    print(a.players[i].name,a.players[i].goals,a.players[i].assists)'''


def league_teams(t):
    for i in range(20):
        print(t[i].team)

def cl_set():
    cl = np.full((4,12),"",dtype=object)
    l = [0,1,2]
    linv = [0,1,2,3,4,5]
    s1 = [[],[],[],[],[],[]]
    '''for i in range(5):
        for j in range(4):
            cl[j,i] = Sheet_teams.cell(3+j,2+3*i).value'''
    for i in range(6):
        for j in range(3):
            s1[i].append(Sheet_teams.cell(3+j,17+3*i).value)
    for i in range(2):
        rl = random.sample(l,2)
        for j in range(2):
            cl[j,5+i] = s1[i][rl[j]]
    for i in range(4):
        rl = random.sample(l,1)
        cl[0,7+i] = s1[2+i][rl[0]]
    s2 = []
    for i in range(6):
        s2.append(Sheet_teams.cell(3+i,35).value)
    rl = random.sample(linv,4)
    for i in range(4):
        cl[i,11] = s2[rl[i]]
    data_cl = pd.DataFrame(cl)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_cl.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=2, startcol=1)

t_all = ts + te + ti + tg + tf + tp + tn + ta + tsc + tr + tb + tinv + ts2
p_all = []
for i in t_all:
    for j in range(len(i.players)):
        p_all.append(i.players[j])

def cl_groupstage(round):
    st = [0]*8
    gt = [0]*32
    sorted_st = [0]*8
    scorers = np.full((213,3),"",dtype=object)
    assisters = np.full((213,3),"",dtype=object)
    for i in range(213):
        scorers[i][0] = p_all[i].name
        scorers[i][1] = 0
        scorers[i][2] = i+1
        assisters[i][0] = p_all[i].name
        assisters[i][1] = 0
        assisters[i][2] = i+1
    scores = [np.full((2,3),"",dtype=object)]*8
    for i in range(8):
        st[i] = np.full((4,11),"",dtype=object)
        sorted_st[i] = np.full((4,11),"",dtype=object)
        if i<4:
            for j in range(4):
                st[i][j,0] = Sheet_cl.cell(15+i*7+j,1).value
                st[i][j,1] = Sheet_cl.cell(15+i*7+j,2).value
                st[i][j,2] = Sheet_cl.cell(15+i*7+j,3).value
                st[i][j,3] = Sheet_cl.cell(15+i*7+j,4).value
                st[i][j,4] = Sheet_cl.cell(15+i*7+j,5).value
                st[i][j,5] = Sheet_cl.cell(15+i*7+j,6).value
                st[i][j,6] = Sheet_cl.cell(15+i*7+j,7).value
                st[i][j,7] = Sheet_cl.cell(15+i*7+j,8).value
                st[i][j,8] = Sheet_cl.cell(15+i*7+j,9).value
                st[i][j,9] = Sheet_cl.cell(15+i*7+j,10).value
                st[i][j,10] = Sheet_cl.cell(15+i*7+j,11).value
        else:
            for j in range(4):
                st[i][j,0] = Sheet_cl.cell(15+(i-4)*7+j,12).value
                st[i][j,1] = Sheet_cl.cell(15+(i-4)*7+j,13).value
                st[i][j,2] = Sheet_cl.cell(15+(i-4)*7+j,14).value
                st[i][j,3] = Sheet_cl.cell(15+(i-4)*7+j,15).value
                st[i][j,4] = Sheet_cl.cell(15+(i-4)*7+j,16).value
                st[i][j,5] = Sheet_cl.cell(15+(i-4)*7+j,17).value
                st[i][j,6] = Sheet_cl.cell(15+(i-4)*7+j,18).value
                st[i][j,7] = Sheet_cl.cell(15+(i-4)*7+j,19).value
                st[i][j,8] = Sheet_cl.cell(15+(i-4)*7+j,20).value
                st[i][j,9] = Sheet_cl.cell(15+(i-4)*7+j,21).value
                st[i][j,10] = Sheet_cl.cell(15+(i-4)*7+j,22).value
    for i in range(8):
        sorted_st[i] = st[i][np.argsort(st[i][:,0])]
        for j in range(4):
            gt[i*4+j]=list(filter(lambda x:x.team==sorted_st[i][j,1],t_all))[0]
    if round == 1:
        print("cl gs round1")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[3+4*i],gt[0+4*i],scorers,assisters)
            scores[i][0,0] = gt[3+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[0+4*i].team
            s = game(gt[1+4*i],gt[2+4*i],scorers,assisters)
            scores[i][1,0] = gt[1+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[2+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=23)
    elif round == 2:
        print("cl gs round2")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[2+4*i],gt[3+4*i],scorers,assisters)
            scores[i][0,0] = gt[2+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[3+4*i].team
            s = game(gt[0+4*i],gt[1+4*i],scorers,assisters)
            scores[i][1,0] = gt[0+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[1+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=27)
    elif round == 3:
        print("cl gs round3")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[2+4*i],gt[0+4*i],scorers,assisters)
            scores[i][0,0] = gt[2+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[0+4*i].team
            s = game(gt[3+4*i],gt[1+4*i],scorers,assisters)
            scores[i][1,0] = gt[3+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[1+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=31)
    elif round == 4:
        print("cl gs round4")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[1+4*i],gt[3+4*i],scorers,assisters)
            scores[i][0,0] = gt[1+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[3+4*i].team
            s = game(gt[0+4*i],gt[2+4*i],scorers,assisters)
            scores[i][1,0] = gt[0+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[2+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=35)
    elif round == 5:
        print("cl gs round5")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[1+4*i],gt[0+4*i],scorers,assisters)
            scores[i][0,0] = gt[1+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[0+4*i].team
            s = game(gt[3+4*i],gt[2+4*i],scorers,assisters)
            scores[i][1,0] = gt[3+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[2+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=39)
    elif round == 6:
        print("cl gs round6")
        for i in range(8):
            if i == 0:
                print("Group A")
            elif i==1:
                print("Group B")
            elif i==2:
                print("Group C")
            elif i==3:
                print("Group D")
            elif i==4:
                print("Group E")
            elif i==5:
                print("Group F")
            elif i==6:
                print("Group G")
            elif i==7:
                print("Group H")
            s = game(gt[2+4*i],gt[1+4*i],scorers,assisters)
            scores[i][0,0] = gt[2+4*i].team
            scores[i][0,1] = str(s[0]) + "-" + str(s[1])
            scores[i][0,2] = gt[1+4*i].team
            s = game(gt[0+4*i],gt[3+4*i],scorers,assisters)
            scores[i][1,0] = gt[0+4*i].team
            scores[i][1,1] = str(s[0]) + "-" + str(s[1])
            scores[i][1,2] = gt[3+4*i].team
            data_scores = pd.DataFrame(scores[i])
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=5+5*i, startcol=43)
    old_scorers = np.full((213,3),"",dtype=object)
    old_assisters = np.full((213,3),"",dtype=object)
    for i in range(213):
        old_scorers[i,0] = Sheet_cl.cell(67+i,3).value
        old_scorers[i,1] = Sheet_cl.cell(67+i,4).value
        old_scorers[i,2] = Sheet_cl.cell(67+i,5).value
        old_assisters[i,0] = Sheet_cl.cell(67+i,7).value
        old_assisters[i,1] = Sheet_cl.cell(67+i,8).value
        old_assisters[i,2] = Sheet_cl.cell(67+i,9).value
    sorted_os = old_scorers[np.argsort(old_scorers[:,2])]
    sorted_oa = old_assisters[np.argsort(old_assisters[:,2])]
    sorted_s = scorers[np.argsort(scorers[:,2])]
    sorted_a = assisters[np.argsort(assisters[:,2])]
    new_s = sorted_os
    new_a = sorted_oa
    for i in range(213):
        if sorted_s[i][1] != None:
            new_s[i][1] += sorted_s[i][1]
        else:
            new_s[i][1] = sorted_s[i][1]
    for i in range(213):
        if sorted_a[i][1] != None:
            new_a[i][1] += sorted_a[i][1]
        else:
            new_a[i][1] = sorted_a[i][1]
    sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
    sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
    data_new_s = pd.DataFrame(sorted_new_s)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_new_s.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=66, startcol=2)
    data_new_a = pd.DataFrame(sorted_new_a)
    with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        data_new_a.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=66, startcol=6)
    new_sorted_st = [0]*8
    for i in range(8):
        for j in range(4):
            if round != 0:
                sorted_st[i][j,3] += 1
            sorted_st[i][j,4] += gt[j+4*i].win
            sorted_st[i][j,5] += gt[j+4*i].draw
            sorted_st[i][j,6] += gt[j+4*i].lose
            sorted_st[i][j,7] += gt[j+4*i].goal
            sorted_st[i][j,8] += gt[j+4*i].goalag
            sorted_st[i][j,9] += gt[j+4*i].goal - gt[j+4*i].goalag
            sorted_st[i][j,10] += 3*gt[j+4*i].win + gt[j+4*i].draw
        new_sorted_st[i] = sorted_st[i][np.argsort(sorted_st[i][:,10])[::-1]]
        data_sorted_st = pd.DataFrame(new_sorted_st[i])
        if i < 4:
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_sorted_st.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=14+7*i, startcol=0)
        else:
            with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                data_sorted_st.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=14+7*(i-4), startcol=11)

pcdr = []
for i in range(20):
    pcdr.append(ts[i])
for i in range(12):
    pcdr.append(ts2[i])
lcdr = [18, 19, 23, 7, 17, 0, 26, 8, 3, 31, 6, 24, 11, 25, 27, 9, 5, 28, 12, 4, 1, 16, 14, 13, 20, 30, 22, 15, 2, 10, 29, 21]
cdr = [0]*32
for i in range(32):
    cdr[i] = pcdr[lcdr[i]]
'''for i in range(32):
    print(cdr[i].team)'''

# cdr16teams = []
# cdr16 = []
# for i in range(16):
#     cdr16.append(list(filter(lambda x:x.team==cdr16teams[i],cdr))[0])
# cdr8teams = []
# cdr8 = []
# for i in range(8):
#     cdr8.append(list(filter(lambda x:x.team==cdr8teams[i],cdr))[0])
# cdr4teams = []
# cdr4 = []
# for i in range(4):
#     cdr4.append(list(filter(lambda x:x.team==cdr4teams[i],cdr))[0])
# cdr2teams = []
# cdr2 = []
# for i in range(2):
#     cdr2.append(list(filter(lambda x:x.team==cdr2teams[i],cdr))[0])

# def cdrgames(round):
#     scorers = np.full((70,3),"",dtype=object)
#     assisters = np.full((70,3),"",dtype=object)
#     for i in range(70):
#         scorers[i][0] = sl.players[i].name
#         scorers[i][1] = 0
#         scorers[i][2] = i+1
#         assisters[i][0] = sl.players[i].name
#         assisters[i][1] = 0
#         assisters[i][2] = i+1
#     if round == 1:
#         scores = np.full((16,3),"",dtype=object)
#         for i in range(16):
#             s = cdrgame(cdr[2*i],cdr[2*i+1],scorers,assisters)
#             scores[i,0] = cdr[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=39, startcol=2)
#     elif round == 2:
#         scores = np.full((16,3),"",dtype=object)
#         for i in range(16):
#             s = game(cdr[2*i+1],cdr[2*i],scorers,assisters)
#             scores[i,0] = cdr[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=39, startcol=6)
#     elif round == 3:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = cdrgame(cdr16[2*i],cdr16[2*i+1],scorers,assisters)
#             scores[i,0] = cdr16[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr16[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=39, startcol=10)
#     elif round == 4:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(cdr16[2*i+1],cdr16[2*i],scorers,assisters)
#             scores[i,0] = cdr16[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr16[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=39, startcol=14)
#     elif round == 5:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = cdrgame(cdr8[2*i],cdr8[2*i+1],scorers,assisters)
#             scores[i,0] = cdr8[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr8[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=50, startcol=10)
#     elif round == 6:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = game(cdr8[2*i+1],cdr8[2*i],scorers,assisters)
#             scores[i,0] = cdr8[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr8[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=50, startcol=14)
#     elif round == 7:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(cdr4[2*i],cdr4[2*i+1],scorers,assisters)
#             scores[i,0] = cdr4[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr4[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=57, startcol=10)
#     elif round == 8:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(cdr4[2*i+1],cdr4[2*i],scorers,assisters)
#             scores[i,0] = cdr4[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr4[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=57, startcol=14)
#     elif round == 9:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(1):
#             s = cdrgame(cdr2[2*i+1],cdr2[2*i],scorers,assisters)
#             scores[i,0] = cdr2[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cdr2[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=61, startcol=10)
#     old_scorers = np.full((70,3),"",dtype=object)
#     old_assisters = np.full((70,3),"",dtype=object)
#     for i in range(70):
#         old_scorers[i,0] = Sheet_cdr.cell(39+i,20).value
#         old_scorers[i,1] = Sheet_cdr.cell(39+i,21).value
#         old_scorers[i,2] = Sheet_cdr.cell(39+i,22).value
#         old_assisters[i,0] = Sheet_cdr.cell(39+i,24).value
#         old_assisters[i,1] = Sheet_cdr.cell(39+i,25).value
#         old_assisters[i,2] = Sheet_cdr.cell(39+i,26).value
#     sorted_os = old_scorers[np.argsort(old_scorers[:,2])]
#     sorted_oa = old_assisters[np.argsort(old_assisters[:,2])]
#     sorted_s = scorers[np.argsort(scorers[:,2])]
#     sorted_a = assisters[np.argsort(assisters[:,2])]
#     new_s = sorted_s
#     new_a = sorted_a
#     for i in range(70):
#         if sorted_os[i][1] != None:
#             new_s[i][1] += sorted_os[i][1]
#         else:
#             new_s[i][1] = sorted_os[i][1]
#     for i in range(70):
#         if sorted_oa[i][1] != None:
#             new_a[i][1] += sorted_oa[i][1]
#         else:
#             new_a[i][1] = sorted_oa[i][1]
#     sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
#     sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
#     data_new_s = pd.DataFrame(sorted_new_s)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_s.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=38, startcol=19)
#     data_new_a = pd.DataFrame(sorted_new_a)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_a.to_excel(writer, sheet_name="cdr",header=False, index=False,startrow=38, startcol=23)

# cl16teams = []
# cl16 = []
# for i in range(16):
#     cl16.append(list(filter(lambda x:x.team==cl16teams[i],t_all))[0])
# cl8teams = []
# cl8 = []
# for i in range(8):
#     cl8.append(list(filter(lambda x:x.team==cl8teams[i],t_all))[0])
# cl4teams = []
# cl4 = []
# for i in range(4):
#     cl4.append(list(filter(lambda x:x.team==cl4teams[i],t_all))[0])
# cl2teams = []
# cl2 = []
# for i in range(2):
#     cl2.append(list(filter(lambda x:x.team==cl2teams[i],t_all))[0])

# def cl_kostage(round):
#     scorers = np.full((213,3),"",dtype=object)
#     assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         scorers[i][0] = p_all[i].name
#         scorers[i][1] = 0
#         scorers[i][2] = i+1
#         assisters[i][0] = p_all[i].name
#         assisters[i][1] = 0
#         assisters[i][2] = i+1
#     if round == 1:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(cl16[2*i],cl16[2*i+1],scorers,assisters)
#             scores[i,0] = cl16[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl16[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=47, startcol=23)
#     elif round == 2:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(cl16[2*i+1],cl16[2*i],scorers,assisters)
#             scores[i,0] = cl16[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl16[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=47, startcol=27)
#     elif round == 3:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = game(cl8[2*i],cl8[2*i+1],scorers,assisters)
#             scores[i,0] = cl8[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl8[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=58, startcol=23)
#     elif round == 4:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = game(cl8[2*i+1],cl8[2*i],scorers,assisters)
#             scores[i,0] = cl8[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl8[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=58, startcol=27)
#     elif round == 5:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(cl4[2*i],cl4[2*i+1],scorers,assisters)
#             scores[i,0] = cl4[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl4[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=65, startcol=23)
#     elif round == 6:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(cl4[2*i+1],cl4[2*i],scorers,assisters)
#             scores[i,0] = cl4[2*i+1].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl4[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=65, startcol=27)
#     elif round == 7:
#         scores = np.full((1,3),"",dtype=object)
#         for i in range(1):
#             s = final_game(cl2[2*i],cl2[2*i+1],scorers,assisters)
#             scores[i,0] = cl2[2*i].team
#             scores[i,1] = str(s[0]) + "-" + str(s[1])
#             scores[i,2] = cl2[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=69, startcol=23)
#     old_scorers = np.full((213,3),"",dtype=object)
#     old_assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         old_scorers[i,0] = Sheet_cl.cell(67+i,3).value
#         old_scorers[i,1] = Sheet_cl.cell(67+i,4).value
#         old_scorers[i,2] = Sheet_cl.cell(67+i,5).value
#         old_assisters[i,0] = Sheet_cl.cell(67+i,7).value
#         old_assisters[i,1] = Sheet_cl.cell(67+i,8).value
#         old_assisters[i,2] = Sheet_cl.cell(67+i,9).value
#     sorted_os = old_scorers[np.argsort(old_scorers[:,2])]
#     sorted_oa = old_assisters[np.argsort(old_assisters[:,2])]
#     sorted_s = scorers[np.argsort(scorers[:,2])]
#     sorted_a = assisters[np.argsort(assisters[:,2])]
#     new_s = sorted_os
#     new_a = sorted_oa
#     for i in range(213):
#         if sorted_s[i][1] != None:
#             new_s[i][1] += sorted_s[i][1]
#         else:
#             new_s[i][1] = sorted_s[i][1]
#     for i in range(213):
#         if sorted_a[i][1] != None:
#             new_a[i][1] += sorted_a[i][1]
#         else:
#             new_a[i][1] = sorted_a[i][1]
#     sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
#     sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
#     data_new_s = pd.DataFrame(sorted_new_s)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_s.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=66, startcol=2)
#     data_new_a = pd.DataFrame(sorted_new_a)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_a.to_excel(writer, sheet_name="cl",header=False, index=False,startrow=66, startcol=6)

# elplayoffteams = []
# elplayoff = []
# for i in range(16):
#     elplayoff.append(list(filter(lambda x:x.team==elplayoffteams[i],t_all))[0])
# el16teams = []
# el16 = []
# for i in range(16):
#     el16.append(list(filter(lambda x:x.team==el16teams[i],t_all))[0])
# el8teams = []
# el8 = []
# for i in range(8):
#     el8.append(list(filter(lambda x:x.team==el8teams[i],t_all))[0])
# el4teams = []
# el4 = []
# for i in range(4):
#     el4.append(list(filter(lambda x:x.team==el4teams[i],t_all))[0])
# el2teams = []
# el2 = []
# for i in range(2):
#     el2.append(list(filter(lambda x:x.team==el2teams[i],t_all))[0])

# def el_playoff(round):
#     scorers = np.full((213,3),"",dtype=object)
#     assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         scorers[i][0] = p_all[i].name
#         scorers[i][1] = 0
#         scorers[i][2] = i+1
#         assisters[i][0] = p_all[i].name
#         assisters[i][1] = 0
#         assisters[i][2] = i+1
#     if round == 1:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(elplayoff[2*i],elplayoff[2*i+1],scorers,assisters)
#             scores[i,0] = elplayoff[2*i].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = elplayoff[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=8, startcol=1)
#     elif round == 2:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(elplayoff[2*i+1],elplayoff[2*i],scorers,assisters)
#             scores[i,0] = elplayoff[2*i+1].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = elplayoff[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=8, startcol=5)
#     old_scorers = np.full((213,3),"",dtype=object)
#     old_assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         old_scorers[i,0] = Sheet_el.cell(44+i,3).value
#         old_scorers[i,1] = Sheet_el.cell(44+i,4).value
#         old_scorers[i,2] = Sheet_el.cell(44+i,5).value
#         old_assisters[i,0] = Sheet_el.cell(44+i,7).value
#         old_assisters[i,1] = Sheet_el.cell(44+i,8).value
#         old_assisters[i,2] = Sheet_el.cell(44+i,9).value
#     sorted_os = old_scorers[np.argsort(old_scorers[:,2])]
#     sorted_oa = old_assisters[np.argsort(old_assisters[:,2])]
#     sorted_s = scorers[np.argsort(scorers[:,2])]
#     sorted_a = assisters[np.argsort(assisters[:,2])]
#     new_s = sorted_os
#     new_a = sorted_oa
#     for i in range(213):
#         if sorted_s[i][1] != None:
#             new_s[i][1] += sorted_s[i][1]
#         else:
#             new_s[i][1] = sorted_s[i][1]
#     for i in range(213):
#         if sorted_a[i][1] != None:
#             new_a[i][1] += sorted_a[i][1]
#         else:
#             new_a[i][1] = sorted_a[i][1]
#     sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
#     sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
#     data_new_s = pd.DataFrame(sorted_new_s)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_s.to_excel(writer, sheet_name="el",header=False, index=False,startrow=43, startcol=2)
#     data_new_a = pd.DataFrame(sorted_new_a)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_a.to_excel(writer, sheet_name="el",header=False, index=False,startrow=43, startcol=6)
# def el_kostage(round):
#     scorers = np.full((213,3),"",dtype=object)
#     assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         scorers[i][0] = p_all[i].name
#         scorers[i][1] = 0
#         scorers[i][2] = i+1
#         assisters[i][0] = p_all[i].name
#         assisters[i][1] = 0
#         assisters[i][2] = i+1
#     if round == 1:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(el16[2*i],el16[2*i+1],scorers,assisters)
#             scores[i,0] = el16[2*i].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el16[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=24, startcol=21)
#     elif round == 2:
#         scores = np.full((8,3),"",dtype=object)
#         for i in range(8):
#             s = game(el16[2*i+1],el16[2*i],scorers,assisters)
#             scores[i,0] = el16[2*i+1].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el16[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=24, startcol=25)
#     elif round == 3:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = game(el8[2*i],el8[2*i+1],scorers,assisters)
#             scores[i,0] = el8[2*i].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el8[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=35, startcol=21)
#     elif round == 4:
#         scores = np.full((4,3),"",dtype=object)
#         for i in range(4):
#             s = game(el8[2*i+1],el8[2*i],scorers,assisters)
#             scores[i,0] = el8[2*i+1].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el8[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=35, startcol=25)
#     elif round == 5:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(el4[2*i],el4[2*i+1],scorers,assisters)
#             scores[i,0] = el4[2*i].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el4[2*i+1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=42, startcol=21)
#     elif round == 6:
#         scores = np.full((2,3),"",dtype=object)
#         for i in range(2):
#             s = game(el4[2*i+1],el4[2*i],scorers,assisters)
#             scores[i,0] = el4[2*i+1].team
#             scores[i,1] = str(s[0])+"-"+str(s[1])
#             scores[i,2] = el4[2*i].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=42, startcol=25)
#     elif round == 7:
#         scores = np.full((1,3),"",dtype=object)
#         s = final_game(el2[0],el2[1],scorers,assisters)
#         scores[0,0] = el2[0].team
#         scores[0,1] = str(s[0])+"-"+str(s[1])
#         scores[0,2] = el2[1].team
#         data_scores = pd.DataFrame(scores)
#         with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             data_scores.to_excel(writer, sheet_name="el",header=False, index=False,startrow=46, startcol=21)
#     old_scorers = np.full((213,3),"",dtype=object)
#     old_assisters = np.full((213,3),"",dtype=object)
#     for i in range(213):
#         old_scorers[i,0] = Sheet_el.cell(44+i,3).value
#         old_scorers[i,1] = Sheet_el.cell(44+i,4).value
#         old_scorers[i,2] = Sheet_el.cell(44+i,5).value
#         old_assisters[i,0] = Sheet_el.cell(44+i,7).value
#         old_assisters[i,1] = Sheet_el.cell(44+i,8).value
#         old_assisters[i,2] = Sheet_el.cell(44+i,9).value
#     sorted_os = old_scorers[np.argsort(old_scorers[:,2])]
#     sorted_oa = old_assisters[np.argsort(old_assisters[:,2])]
#     sorted_s = scorers[np.argsort(scorers[:,2])]
#     sorted_a = assisters[np.argsort(assisters[:,2])]
#     new_s = sorted_os
#     new_a = sorted_oa
#     for i in range(213):
#         if sorted_s[i][1] != None:
#             new_s[i][1] += sorted_s[i][1]
#         else:
#             new_s[i][1] = sorted_s[i][1]
#     for i in range(213):
#         if sorted_a[i][1] != None:
#             new_a[i][1] += sorted_a[i][1]
#         else:
#             new_a[i][1] = sorted_a[i][1]
#     sorted_new_s = new_s[np.argsort(new_s[:,1])[::-1]]
#     sorted_new_a = new_a[np.argsort(new_a[:,1])[::-1]]
#     data_new_s = pd.DataFrame(sorted_new_s)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_s.to_excel(writer, sheet_name="el",header=False, index=False,startrow=43, startcol=2)
#     data_new_a = pd.DataFrame(sorted_new_a)
#     with pd.ExcelWriter("fb_real.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         data_new_a.to_excel(writer, sheet_name="el",header=False, index=False,startrow=43, startcol=6)

def cup_game(a,b,l):
    s = score(a.power,b.power)
    while s[0]==s[1]:
        s = score(a.power,b.power)
    if s[0] > s[1]:
        l.append(a)
    elif s[0] < s[1]:
        l.append(b)
    print(a.team+" "+str(s[0])+":"+str(s[1])+" "+b.team)
    print(str(s[2])+"|"+str(s[3]))
    print("")

# eteams16 = []
# ec16 = []
# for i in range(len(eteams16)):
#     ec16.append(list(filter(lambda x:x.team==eteams16[i],te))[0])
# eteams8 = []
# ec8 = []
# for i in range(len(eteams8)):
#     ec8.append(list(filter(lambda x:x.team==eteams8[i],te))[0])
# eteams4 = []
# ec4 = []
# for i in range(len(eteams4)):
#     ec4.append(list(filter(lambda x:x.team==eteams4[i],te))[0])
# eteams2 = []
# ec2 = []
# for i in range(len(eteams2)):
#     ec2.append(list(filter(lambda x:x.team==eteams2[i],te))[0])
# ewinner =[]

# iteams16 = []
# ic16 = []
# for i in range(len(iteams16)):
#     ic16.append(list(filter(lambda x:x.team==iteams16[i],ti))[0])
# iteams8 = []
# ic8 = []
# for i in range(len(iteams8)):
#     ic8.append(list(filter(lambda x:x.team==iteams8[i],ti))[0])
# iteams4 = []
# ic4 = []
# for i in range(len(iteams4)):
#     ic4.append(list(filter(lambda x:x.team==iteams4[i],ti))[0])
# iteams2 = []
# ic2 = []
# for i in range(len(iteams2)):
#     ic2.append(list(filter(lambda x:x.team==iteams2[i],ti))[0])
# iwinner =[]

# gteams16 = []
# gc16 = []
# for i in range(len(gteams16)):
#     gc16.append(list(filter(lambda x:x.team==gteams16[i],tg))[0])
# gteams8 = []
# gc8 = []
# for i in range(len(gteams8)):
#     gc8.append(list(filter(lambda x:x.team==gteams8[i],tg))[0])
# gteams4 = []
# gc4 = []
# for i in range(len(gteams4)):
#     gc4.append(list(filter(lambda x:x.team==gteams4[i],tg))[0])
# gteams2 = []
# gc2 = []
# for i in range(len(gteams2)):
#     gc2.append(list(filter(lambda x:x.team==gteams2[i],tg))[0])
# gwinner =[]

# fteams16 = []
# fc16 = []
# for i in range(len(fteams16)):
#     fc16.append(list(filter(lambda x:x.team==fteams16[i],tf))[0])
# fteams8 = []
# fc8 = []
# for i in range(len(fteams8)):
#     fc8.append(list(filter(lambda x:x.team==fteams8[i],tf))[0])
# fteams4 = []
# fc4 = []
# for i in range(len(fteams4)):
#     fc4.append(list(filter(lambda x:x.team==fteams4[i],tf))[0])
# fteams2 = []
# fc2 = []
# for i in range(len(fteams2)):
#     fc2.append(list(filter(lambda x:x.team==fteams2[i],tf))[0])
# fwinner =[]

def cup(t,n,tb,ta):
    ta = []
    if n==1:
        l = list(range(10))
        t16 = []
        random.shuffle(l)
        for i in range(10):
            t16.append(t[i])
        for i in range(6):
            t16.append(t[10+l[i]])
        random.shuffle(t16)
        t16teams = []
        for i in range(16):
            print(t16[i].team)
        for i in range(16):
            t16teams.append(t16[i].team)
        #print(t16teams)
        print("")
        return t16
    if n==2:
        for i in range(8):
            cup_game(tb[2*i],tb[2*i+1],ta)
        t8teams = []
        for i in range(8):
            print(ta[i].team)
        for i in range(8):
            t8teams.append(ta[i].team)
        #print(t8teams)
        print("")
        print("")
        return ta
    if n==3:
        for i in range(4):
            cup_game(tb[2*i],tb[2*i+1],ta)
        t4teams = []
        for i in range(4):
            print(ta[i].team)
        for i in range(4):
            t4teams.append(ta[i].team)
        #print(t4teams)
        print("")
        print("")
        return ta
    if n==4:
        for i in range(2):
            cup_game(tb[2*i],tb[2*i+1],ta)
        t2teams = []
        for i in range(2):
            print(ta[i].team)  
        for i in range(2):
            t2teams.append(ta[i].team)
        #print(t2teams)
        print("")
        print("")
        return ta
    if n==5:
        cup_game(tb[0],tb[1],ta)
        print(ta[0].team)

def cupresult(t):
    tb = []
    ta = []
    tb16 = cup(t,1,tb,ta)
    ta=[]
    tb8 = cup(t,2,tb16,ta)
    ta=[]
    tb4 = cup(t,3,tb8,ta)
    ta = []
    tb2 = cup(t,4,tb4,ta)
    ta = []
    tb1 = cup(t,5,tb2,ta)

sl2_poteams = []
def sl_playoff(t):
    po4 = []
    po2 = []
    w = []
    for i in range(4):
        po4.append(list(filter(lambda x:x.team==t[i],t_all))[0])
    cup_game(po4[0],po4[3],po2)
    cup_game(po4[1],po4[2],po2)
    cup_game(po2[0],po2[1],w)
    for i in range(4):
        print(po4[i].team)
    print(" ") 
    for i in range(2):
        print(po2[i].team)
    print(" ")
    print(w[0].team)

def el_group(n):
    for i in range(3):
        print(score(n+500,3000))
    for i in range(3):
        print(score(n,3500))

#sl_playoff(sl2_poteams)

'''cupresult(te)
cupresult(ti)
cupresult(tg)
cupresult(tf)'''

#el_group(5000)
#el_playoff(2)
#el_kostage(7)

#cl_set()
#cl_groupstage(6)
#cl_kostage(7)

#all_leagueround(38)
gah=[[],[],
[],[]]
gaa=[[],[],
[],[]]
#my_leagueround(37,gah)
#my_leagueround(38,gaa)

#cdrgames(9)

#print(score(4500,4000))